"""
scn.py — Single-Column Notation (SCN) parser and serializer.

A structured data format designed for single-column entry in spreadsheets.
One value per cell, with sections, key/value pairs, lists, dict lists,
and comments. All values are returned as strings.

Public API:
    parse(cells)       — general-purpose SCN parser
    parse_entry(cells) — data-entry variant (consumes next cell as value)
    serialize(data)    — dict → SCN lines
    read_excel(path)   — read .xlsx column → parse()
    read_text(path)    — read .txt → parse()
"""

from __future__ import annotations

from typing import Any


def _set_nested(d: dict, dotted_key: str, value: Any) -> None:
    """Set a value in a nested dict via dot-separated key path.

    Args:
        d: Target dict.
        dotted_key: Dot-separated key (e.g., "parent.child").
        value: Value to set.
    """
    parts = dotted_key.split(".")
    for part in parts[:-1]:
        if part not in d or not isinstance(d[part], dict):
            d[part] = {}
        d = d[part]
    d[parts[-1]] = value


def _get_nested(d: dict, dotted_key: str) -> Any:
    """Get a value from a nested dict via dot-separated key path.

    Args:
        d: Source dict.
        dotted_key: Dot-separated key (e.g., "parent.child").

    Returns:
        The value, or None if path does not exist.
    """
    parts = dotted_key.split(".")
    for part in parts[:-1]:
        if not isinstance(d, dict) or part not in d:
            return None
        d = d[part]
    if not isinstance(d, dict):
        return None
    return d.get(parts[-1])


def _get_nested_parent(d: dict, dotted_key: str) -> tuple[dict, str]:
    """Navigate dot-separated key, return (parent_dict, final_key).

    Args:
        d: Root dict.
        dotted_key: Dot-separated key path.

    Returns:
        Tuple of (parent dict, final key segment).
    """
    parts = dotted_key.split(".")
    for part in parts[:-1]:
        if part not in d or not isinstance(d[part], dict):
            d[part] = {}
        d = d[part]
    return d, parts[-1]


# ---------------------------------------------------------------------------
# General-purpose parser
# ---------------------------------------------------------------------------


def parse(cells: list[Any]) -> dict:
    """Parse a list of SCN values into a nested dict.

    Processes cells top-to-bottom, recognizing all six SCN constructs:
    [section], key:, - item, +name, ;; comment, and empty rows.

    Args:
        cells: Raw values from a spreadsheet column or text file lines.

    Returns:
        Parsed dict. All leaf values are strings.
    """
    root: dict = {}
    stack: list[dict] = [root]
    list_registry: dict[tuple[int, str], list] = {}
    pending_key: str | None = None
    pending_is_list = False

    def current_dict() -> dict:
        return stack[-1]

    def find_list_owner(name: str) -> tuple[int | None, list | None]:
        for si in range(len(stack) - 1, -1, -1):
            key = (id(stack[si]), name)
            if key in list_registry:
                return si, list_registry[key]
        return None, None

    lines = [str(c).strip() if c is not None else "" for c in cells]

    for line in lines:
        if not line:
            continue

        # Plain value — checked FIRST so values starting with ;;, [, +, -
        # are not misinterpreted as constructs
        if pending_key is not None and not pending_is_list:
            if not line.startswith("- "):
                _set_nested(current_dict(), pending_key, line)
                pending_key = None
                continue

        # Comment
        if line.startswith(";;"):
            continue

        # [section]
        if line.startswith("[") and line.endswith("]"):
            section_name = line[1:-1].strip()
            if section_name not in root:
                root[section_name] = {}
            stack.clear()
            stack.append(root)
            stack.append(root[section_name])
            pending_key = None
            pending_is_list = False
            list_registry.clear()
            continue

        # +name (dict list entry)
        if line.startswith("+") and not line.startswith("+ "):
            list_name = line[1:].strip()
            pending_key = None
            pending_is_list = False

            owner_si, lst = find_list_owner(list_name)
            if lst is not None:
                del stack[owner_si + 1 :]
                new_dict: dict = {}
                lst.append(new_dict)
                stack.append(new_dict)
            else:
                ctx = current_dict()
                parent, final_key = _get_nested_parent(ctx, list_name)
                new_list: list = []
                parent[final_key] = new_list
                new_dict = {}
                new_list.append(new_dict)
                list_registry[(id(ctx), list_name)] = new_list
                stack.append(new_dict)
            continue

        # - item (list item)
        if line.startswith("- "):
            item_value = line[2:].strip()
            if pending_key is not None:
                ctx = current_dict()
                parent, final_key = _get_nested_parent(ctx, pending_key)
                if final_key not in parent or not isinstance(parent[final_key], list):
                    parent[final_key] = []
                parent[final_key].append(item_value)
                pending_is_list = True
            continue

        # key: declaration
        if line.endswith(":"):
            pending_key = line[:-1].strip()
            pending_is_list = False
            continue

    return root


# ---------------------------------------------------------------------------
# Data-entry parser
# ---------------------------------------------------------------------------


def parse_entry(cells: list[Any]) -> dict:
    """Parse SCN data entry format where key: always consumes the next cell.

    Unlike parse(), this always treats the cell immediately following a
    key: declaration as its value, even if that cell is empty. Designed
    for Excel data entry sheets where unfilled fields are empty cells.

    Supports sections, key/value, lists, dict lists (+name), and comments.

    Args:
        cells: Raw values from a spreadsheet column.

    Returns:
        Parsed dict. Sections become top-level dict keys.
    """
    root: dict = {}
    section: dict = root
    current: dict = root
    dict_lists: dict[str, list] = {}

    i = 0
    n = len(cells)
    normalized = [str(c).strip() if c is not None else "" for c in cells]

    while i < n:
        line = normalized[i]
        i += 1

        if not line:
            continue

        # Comment
        if line.startswith(";;"):
            continue

        # [section]
        if line.startswith("[") and line.endswith("]"):
            name = line[1:-1].strip()
            if name not in root:
                root[name] = {}
            section = root[name]
            current = section
            dict_lists.clear()
            continue

        # +name (dict list entry)
        if line.startswith("+") and not line.startswith("+ "):
            list_name = line[1:].strip()
            if list_name not in dict_lists:
                section[list_name] = []
                dict_lists[list_name] = section[list_name]
            new_entry: dict = {}
            dict_lists[list_name].append(new_entry)
            current = new_entry
            continue

        # key: declaration — consume next cell as value
        if line.endswith(":"):
            key = line[:-1].strip()
            if i < n:
                next_line = normalized[i]
                if next_line.startswith("- "):
                    # List value: collect consecutive - items
                    items: list[str] = []
                    while i < n and normalized[i].startswith("- "):
                        items.append(normalized[i][2:].strip())
                        i += 1
                    _set_nested(current, key, items)
                else:
                    # Scalar: consume next cell (even if empty)
                    i += 1
                    if next_line:
                        _set_nested(current, key, next_line)
            continue

    return root


# ---------------------------------------------------------------------------
# Serializer
# ---------------------------------------------------------------------------


def _emit_dict(d: dict, prefix: str, lines: list[str]) -> None:
    """Serialize dict contents using dot notation for nested dicts."""
    for key, value in d.items():
        full_key = f"{prefix}.{key}" if prefix else key
        if isinstance(value, dict):
            _emit_dict(value, full_key, lines)
        elif isinstance(value, list):
            if value and isinstance(value[0], dict):
                for entry in value:
                    lines.append(f"+{full_key}")
                    _emit_dict(entry, "", lines)
            else:
                lines.append(f"{full_key}:")
                for item in value:
                    lines.append(f"- {item}")
        else:
            lines.append(f"{full_key}:")
            lines.append(str(value) if value is not None else "")


def serialize(data: dict) -> list[str]:
    """Convert a dict to SCN lines.

    Top-level dict values are rendered as [section] blocks.
    Nested dicts use dot notation. Lists use - items or +name entries.

    Args:
        data: The dict to serialize.

    Returns:
        List of SCN lines (strings).
    """
    lines: list[str] = []

    # Root-level non-dict entries first
    for key, value in data.items():
        if isinstance(value, dict):
            continue
        if isinstance(value, list):
            if value and isinstance(value[0], dict):
                for entry in value:
                    lines.append(f"+{key}")
                    _emit_dict(entry, "", lines)
            else:
                lines.append(f"{key}:")
                for item in value:
                    lines.append(f"- {item}")
        else:
            lines.append(f"{key}:")
            lines.append(str(value) if value is not None else "")

    # Sections (top-level dict values)
    for key, value in data.items():
        if not isinstance(value, dict):
            continue
        if lines and lines[-1] != "":
            lines.append("")
        lines.append(f"[{key}]")
        _emit_dict(value, "", lines)

    # Strip trailing empty lines
    while lines and lines[-1] == "":
        lines.pop()

    return lines


# ---------------------------------------------------------------------------
# Convenience readers
# ---------------------------------------------------------------------------


def read_excel(path: str, sheet: str | None = None, column: int = 1) -> dict:
    """Read a single column from an .xlsx file and parse it.

    Args:
        path: Path to the .xlsx file.
        sheet: Sheet name (defaults to active sheet).
        column: Column number, 1-based (default 1).

    Returns:
        Parsed dict from the column data.
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    cells = [ws.cell(row=r, column=column).value for r in range(1, ws.max_row + 1)]
    wb.close()
    return parse(cells)


def read_text(path: str) -> dict:
    """Read a text file (one entry per line) and parse it.

    Args:
        path: Path to the text file.

    Returns:
        Parsed dict from the file lines.
    """
    with open(path, encoding="utf-8") as f:
        cells = f.read().splitlines()
    return parse(cells)
