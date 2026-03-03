"""
Single-Column Notation (SCN) parser.

A structured data format designed for single-column entry in spreadsheets.

Constructs:
    [section]       top-level named section
    key:            key declaration (value on next line)
    nested.key:     dot-notation nesting (value on next line)
    - item          list item (appends to current key's list)
    +name           new dict entry in list "name"
    ;; comment      ignored (;; must be the first characters)
    (empty rows)    ignored

All values are returned as strings. Type casting and validation
are the responsibility of a separate layer.

Usage:
    from scn import parse, read_excel, read_text

    data = parse(["[section]", "key:", "value"])
    data = read_excel("config.xlsx")
    data = read_text("config.txt")
"""


def _set_nested(d: dict, dotted_key: str, value: str) -> None:
    """Set value in a nested dict via dot-separated key."""
    parts = dotted_key.split(".")
    for part in parts[:-1]:
        if part not in d or not isinstance(d[part], dict):
            d[part] = {}
        d = d[part]
    d[parts[-1]] = value


def _get_nested_parent(d: dict, dotted_key: str):
    """Navigate dot-separated key, return (parent_dict, final_key)."""
    parts = dotted_key.split(".")
    for part in parts[:-1]:
        if part not in d or not isinstance(d[part], dict):
            d[part] = {}
        d = d[part]
    return d, parts[-1]


def parse(cells: list) -> dict:
    """
    Parse a list of values into a nested dict.

    Parameters
    ----------
    cells : list
        Raw values read top-to-bottom (from a column, text lines, etc).

    Returns
    -------
    dict
        Parsed structured data. All values are strings.
    """
    root = {}
    stack = [root]
    list_registry = {}
    pending_key = None
    pending_is_list = False

    def current_dict():
        return stack[-1]

    def find_list_owner(name):
        for si in range(len(stack) - 1, -1, -1):
            key = (id(stack[si]), name)
            if key in list_registry:
                return si, list_registry[key]
        return None, None

    lines = [str(c).strip() if c is not None else "" for c in cells]

    for line in lines:
        if not line:
            continue

        # plain value (follows a key:) — checked FIRST so that values
        # starting with ;;, [, +, or - are not misinterpreted as constructs
        if pending_key is not None and not pending_is_list:
            if not line.startswith("- "):
                _set_nested(current_dict(), pending_key, line)
                pending_key = None
                continue

        # comment
        if line.startswith(";;"):
            continue

        # [section]
        if line.startswith("[") and line.endswith("]"):
            section_name = line[1:-1].strip()
            if section_name not in root:
                root[section_name] = {}
            stack.clear()
            stack.append(root)
            stack.append(root[section_name])
            pending_key = None
            pending_is_list = False
            list_registry.clear()
            continue

        # +name
        if line.startswith("+") and not line.startswith("+ "):
            list_name = line[1:].strip()
            pending_key = None
            pending_is_list = False

            owner_si, lst = find_list_owner(list_name)
            if lst is not None:
                del stack[owner_si + 1:]
                new_dict = {}
                lst.append(new_dict)
                stack.append(new_dict)
            else:
                ctx = current_dict()
                parent, final_key = _get_nested_parent(ctx, list_name)
                new_list = []
                parent[final_key] = new_list
                new_dict = {}
                new_list.append(new_dict)
                list_registry[(id(ctx), list_name)] = new_list
                stack.append(new_dict)
            continue

        # - item
        if line.startswith("- "):
            item_value = line[2:].strip()
            if pending_key is not None:
                ctx = current_dict()
                parent, final_key = _get_nested_parent(ctx, pending_key)
                if final_key not in parent or not isinstance(parent[final_key], list):
                    parent[final_key] = []
                parent[final_key].append(item_value)
                pending_is_list = True
            continue

        # key:
        if line.endswith(":"):
            pending_key = line[:-1].strip()
            pending_is_list = False
            continue



    return root


# ---------------------------------------------------------------------------
# Convenience helpers
# ---------------------------------------------------------------------------

def read_excel(path: str, sheet=None, column: int = 1) -> dict:
    """Read a single column from an .xlsx file and parse it."""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    cells = [ws.cell(row=r, column=column).value for r in range(1, ws.max_row + 1)]
    wb.close()
    return parse(cells)


def read_text(path: str) -> dict:
    """Read a .txt file (one entry per line) and parse it."""
    with open(path, "r", encoding="utf-8") as f:
        cells = f.read().splitlines()
    return parse(cells)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import json, sys

    if len(sys.argv) < 2:
        print("Usage: python scn.py <file.xlsx|file.txt> [sheet] [column]")
        sys.exit(1)

    path = sys.argv[1]
    if path.endswith(".xlsx") or path.endswith(".xlsm"):
        sheet = sys.argv[2] if len(sys.argv) > 2 else None
        col = int(sys.argv[3]) if len(sys.argv) > 3 else 1
        result = read_excel(path, sheet=sheet, column=col)
    else:
        result = read_text(path)

    print(json.dumps(result, indent=2, default=str))
